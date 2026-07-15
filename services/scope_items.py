"""
Catálogo de Scope Items de SAP Best Practices para SAP S/4HANA Cloud.

FUENTE DE VERDAD: la hoja 'S0' del archivo 'Estimador S0 V2.0.xlsx' de SEIDOR,
cuya columna 'SAP S/4HANA BP' contiene los IDs oficiales de scope items por
proceso de negocio. Este módulo la lee dinámicamente (con caché por mtime),
de modo que si los consultores actualizan el Estimador, la lámina de alcance
del PPT se actualiza sola.

Si el Excel no está disponible o la hoja cambia de estructura, se usa el
catálogo estático de respaldo (derivado de la misma hoja S0 V2.0).
"""

import os
import re
import threading
import logging

log = logging.getLogger("scope_items")

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "Estimador S0 V2.0.xlsx")
S0_SHEET = "S0"

# Ediciones soportadas de SAP S/4HANA Cloud
EDITION_PUBLIC = "Public"
EDITION_PRIVATE = "Private"
VALID_EDITIONS = (EDITION_PUBLIC, EDITION_PRIVATE)

EDITION_LABELS = {
    EDITION_PUBLIC: {
        "nombre": "SAP S/4HANA Cloud, Public Edition",
        "programa": "GROW with SAP",
        "descripcion": "Nube pública multi-tenant con mejores prácticas preconfiguradas, "
                       "actualizaciones automáticas semestrales y time-to-value acelerado.",
    },
    EDITION_PRIVATE: {
        "nombre": "SAP S/4HANA Cloud, Private Edition",
        "programa": "RISE with SAP",
        "descripcion": "Instancia dedicada en la nube con máxima flexibilidad de extensión, "
                       "ideal para migraciones desde SAP ECC con desarrollos a medida.",
    },
}

# Orden canónico de módulos en la lámina de alcance
MODULE_ORDER = ("FI", "CO", "MM", "SD", "PP", "PS", "EWM", "QM", "PM")

# Respaldo estático derivado de la hoja S0 del Estimador S0 V2.0 (columna 'SAP S/4HANA BP').
# Cada entrada es (scope_item_id | None, nombre del proceso). None = proceso del
# Estimador sin scope item estándar asociado (se implementa vía Fit-to-Standard).
FALLBACK_SCOPE_CATALOG = {
    "FI": [
        ("J58", "Contabilidad General y Cierre Contable"),
        ("J59", "Cuentas por Cobrar"),
        ("J60", "Cuentas por Pagar y Pago a Proveedores"),
        ("J62", "Gestión de Activos Fijos"),
        ("BFA", "Maestro y Gestión de Bancos"),
        ("4X8", "Gestión de Tesorería"),
        ("BD6", "Gestión de Créditos"),
    ],
    "CO": [
        ("J54", "Control de Ingresos, Costos y Gastos"),
        ("J55", "Análisis de Rentabilidad (CO-PA)"),
        ("1HB", "Planificación de Ingresos, Costos y Gastos"),
        ("4RC", "Reportes de Análisis"),
        ("BEV", "Cierre de Controlling"),
    ],
    "MM": [
        ("J45", "Compras de Stock"),
        ("22Z", "Compra de Servicios"),
        ("BNX", "Compras de Consumibles"),
        ("18J", "Solicitudes de Compra Imputadas"),
        ("BMD", "Compras con Contratos"),
        ("2LG", "Consignación de Proveedores"),
        ("BMY", "Subcontratación"),
        ("BMC", "Gestión de Stocks"),
        ("4LU", "Gestión de Inventarios"),
    ],
    "SD": [
        ("BD9", "Venta desde Stock"),
        ("BKA", "Precios y Condiciones de Venta"),
        ("BKJ", "Venta con Anticipos"),
        ("BD3", "Venta Calzada (Back-to-Back)"),
        ("BDA", "Transferencia Gratuita y Donaciones"),
        ("5D2", "Venta Intercompañía Avanzada"),
        ("1IU", "Consignación de Clientes"),
        ("I9I", "Contratos por Valor y Cantidad"),
    ],
    "PP": [
        ("BJ5", "Fabricación contra Stock – Discreta / MPS"),
        ("J44", "Planificación de Necesidades (MRP)"),
        ("BJE", "Fabricación contra Pedido – Discreta"),
        ("BJ8", "Fabricación contra Stock – Procesos"),
        ("BJH", "Fabricación Repetitiva"),
        ("BJK", "Fabricación con Subcontratación"),
    ],
    "PS": [
        (None, "Datos Maestros de Proyectos (WBS/PEP)"),
        (None, "Planificación Detallada del Proyecto"),
        (None, "Presupuesto de Proyectos y Disponibilidad"),
        (None, "Ejecución e Imputaciones a Proyectos"),
        (None, "Cierre Técnico y Liquidación Final"),
        (None, "Reportes de Proyectos"),
    ],
    "EWM": [
        ("1FS", "Procesos de Entrada a Almacén"),
        ("1G2", "Procesos de Salida de Almacén"),
        ("1V5", "Procesos Internos de Almacén"),
        ("1V7", "Inventario en Almacén"),
    ],
    "QM": [
        ("1FM", "Inspección de Calidad en Abastecimiento"),
        ("1E1", "Gestión de Calidad en Producción"),
        ("1MR", "Gestión de Calidad en Almacenes"),
        ("2F9", "Gestión de Reclamos a Proveedores"),
    ],
    "PM": [
        ("BH1", "Mantenimiento Correctivo"),
        ("BJ2", "Mantenimiento Preventivo"),
    ],
}

# Caché de lectura del Excel (mismo patrón que financial_engine)
_lock = threading.Lock()
_s0_cache = None
_s0_last_mtime = 0


def normalize_edition(value):
    """
    Normaliza cualquier variante ('private', 'Privada', 'RISE', etc.) a
    'Public' o 'Private'. Ante valor desconocido retorna 'Public' (default GROW).
    """
    if isinstance(value, str):
        v = value.strip().lower()
        if v.startswith("priv") or "rise" in v:
            return EDITION_PRIVATE
    return EDITION_PUBLIC


def _clean_process_name(raw):
    """'FI-05_Cuentas por Pagar(FI_P020)' -> 'Cuentas por Pagar'"""
    name = str(raw)
    if "_" in name:
        name = name.split("_", 1)[1]
    name = re.sub(r"\([^)]*\)\s*$", "", name)
    return re.sub(r"\s+", " ", name).strip()


def _extract_base_ids(bp_value):
    """
    Extrae los IDs base de scope item de la celda 'SAP S/4HANA BP'.
    Maneja separadores por salto de línea, '/', espacios, y variantes con
    sufijo ('J60-05' -> 'J60', 'J58-00_J58-02' -> 'J58').
    """
    if bp_value is None:
        return []
    ids = []
    for token in re.split(r"[\s/,]+", str(bp_value).strip()):
        token = token.strip().upper()
        if not token or token == "-":
            continue
        base = token.split("-")[0].split("_")[0]
        if 2 <= len(base) <= 4 and re.match(r"^[0-9A-Z]+$", base) and base not in ids:
            ids.append(base)
    return ids


def _parse_s0_sheet():
    """Lee la hoja S0 del Estimador y agrupa scope items por módulo."""
    import openpyxl
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    try:
        if S0_SHEET not in wb.sheetnames:
            raise ValueError(f"La hoja '{S0_SHEET}' no existe en el Estimador.")
        ws = wb[S0_SHEET]
        catalog = {}
        seen_ids = {}
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            proceso = row[2] if len(row) > 2 else None
            bp = row[3] if len(row) > 3 else None
            if not proceso or "-" not in str(proceso):
                continue
            module = str(proceso).split("-")[0].strip().upper()
            if module not in MODULE_ORDER:
                continue
            name = _clean_process_name(proceso)
            # Omitir filas de estructura organizativa y datos maestros genéricos
            if name.lower().startswith(("estructura organizativa", "datos maestros de mantenimiento")):
                continue
            ids = _extract_base_ids(bp)
            bucket = catalog.setdefault(module, [])
            seen = seen_ids.setdefault(module, set())
            if ids:
                for sid in ids:
                    if sid not in seen:
                        seen.add(sid)
                        bucket.append((sid, name))
            elif module == "PS":
                # PS no tiene scope items estándar en el Estimador: se listan procesos
                bucket.append((None, name))
        if not catalog:
            raise ValueError("La hoja S0 no produjo ningún scope item.")
        return catalog
    finally:
        wb.close()


def load_scope_catalog():
    """
    Retorna el catálogo {módulo: [(id|None, nombre), ...]} leído del Estimador,
    con caché por mtime. Ante cualquier error usa el respaldo estático.
    """
    global _s0_cache, _s0_last_mtime
    try:
        mtime = os.path.getmtime(EXCEL_PATH)
        with _lock:
            if _s0_cache is not None and mtime == _s0_last_mtime:
                return _s0_cache
        parsed = _parse_s0_sheet()
        with _lock:
            _s0_cache = parsed
            _s0_last_mtime = mtime
        log.info("[SCOPE] Catálogo de scope items cargado desde la hoja S0 del Estimador.")
        return parsed
    except Exception as e:
        log.warning("[SCOPE] No se pudo leer la hoja S0 (%s). Usando catálogo de respaldo.", e)
        return FALLBACK_SCOPE_CATALOG


def get_scope_items(active_modules):
    """
    Retorna el subconjunto ordenado del catálogo para los módulos activos.

    Parámetros:
    - active_modules (iterable): módulos SAP activos, ej. ['FI', 'CO', 'MM'].

    Retorna:
    - list[tuple]: lista de (módulo, [(id|None, nombre), ...]) en orden canónico.
    """
    catalog = load_scope_catalog()
    result = []
    for mod in MODULE_ORDER:
        if mod in active_modules and catalog.get(mod):
            result.append((mod, catalog[mod]))
    return result
