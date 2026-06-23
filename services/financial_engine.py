import openpyxl
import os
import sqlite3
import shutil
import tempfile
import threading
import logging

log = logging.getLogger("financial_engine")

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "Estimador S0 V2.0.xlsx")
DB_NAME = os.getenv("DB_NAME", "proposals.db")

DEFAULT_MODULAR_LICENSES = {
    'FI': 20000,
    'CO': 15000,
    'MM': 15000,
    'SD': 15000,
    'PP': 20000,
    'PS': 15000,
    'EWM': 15000,
    'QM': 10000,
    'PM': 10000
}

# Variables globales para el almacenamiento en caché de Excel
_lock = threading.Lock()
_excel_cache = {}
_excel_last_mtime = 0

def format_usd(val):
    """
    Formatea un valor flotante en el formato estándar de Dólares Americanos (USD).
    Ejemplo: $45,000.00
    """
    return f"${val:,.2f}"

def format_pen(val):
    """
    Formatea un valor flotante en el formato estándar de Soles Peruanos (PEN).
    Ejemplo: S/. 150,000.00
    """
    return f"S/. {val:,.2f}"

def load_db_config():
    """
    Carga los parámetros comerciales en tiempo real desde la tabla 'configuracion_comercial' de SQLite.
    Retorna un diccionario con las tarifas base configuradas y parámetros de localización.
    """
    config = {
        'tarifa_hora_consultor': 60.00,
        'porcentaje_ams': 0.15,
        'margen_saas': 0.20,
        'anos_roi': 5.0,
        'factor_igv': 0.18,
        'tipo_cambio_pen': 3.78
    }
    
    if os.path.exists(DB_NAME):
        conn = sqlite3.connect(DB_NAME)
        try:
            cursor = conn.cursor()
            cursor.execute("SELECT parametro, valor FROM configuracion_comercial")
            rows = cursor.fetchall()
            for parametro, valor in rows:
                if parametro in config:
                    config[parametro] = float(valor)
        except sqlite3.Error as e:
            log.warning("No se pudo leer configuracion_comercial de BD: %s. Usando defaults.", e)
        finally:
            conn.close()
    return config

def _load_workbook_safe():
    """Carga el workbook con hot-copy fallback para archivos bloqueados en Windows."""
    try:
        return openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    except (PermissionError, OSError) as e:
        log.warning("[CACHE] Error al abrir Excel directamente (bloqueado): %s. Copiando en caliente...", e)
        temp_dir = os.path.join(tempfile.gettempdir(), "grow_deck_automator")
        os.makedirs(temp_dir, exist_ok=True)
        temp_path = os.path.join(temp_dir, "temp_estimador.xlsx")
        try:
            shutil.copy2(EXCEL_PATH, temp_path)
            return openpyxl.load_workbook(temp_path, data_only=True)
        except Exception as copy_err:
            raise OSError(f"Error comercial: El archivo de estimaciones está bloqueado o inaccesible (incluso en copia temporal): {copy_err}")


def _find_prom_row(sheet):
    """Localiza la fila donde comienza la sección PROM."""
    for r in range(1, 100):
        if sheet.cell(r, 1).value == 'PROM':
            return r
    raise ValueError("Error de integridad: No se encontró la sección de promedios 'PROM' en la hoja Estimar.")


def _parse_excel_row(r, sheet):
    """Parsea una fila del Excel extrayendo semanas y asignación FTE con tolerancia a fallos."""
    try:
        exp_weeks = float(sheet.cell(r, 2).value or 0.0)
        real_weeks = float(sheet.cell(r, 3).value or 0.0)
        deploy_weeks = float(sheet.cell(r, 4).value or 0.0)
        ttgl_weeks = float(sheet.cell(r, 5).value or 0.0)
        asig_real = float(sheet.cell(r, 7).value or 0.0)

        if ttgl_weeks <= 0.0:
            ttgl_weeks = exp_weeks + real_weeks + deploy_weeks
        if ttgl_weeks <= 0.0:
            ttgl_weeks = 20.0
        if asig_real <= 0.0:
            asig_real = 0.25
        return exp_weeks, real_weeks, deploy_weeks, ttgl_weeks, asig_real
    except (ValueError, TypeError):
        log.warning("[WARNING] Datos corruptos en fila %s. Aplicando fallback de seguridad.", r)
        return 4.0, 12.0, 4.0, 20.0, 0.5


def load_excel_data_cached():
    """
    Carga los datos del archivo de presupuesto Excel con un mecanismo de caché
    en memoria que evita re-lecturas innecesarias si el archivo no
    sufre modificaciones (mtime).

    Aplica controles estrictos de integridad de datos y tolerancia a fallos.
    """
    global _excel_cache, _excel_last_mtime

    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Error comercial: El archivo de presupuesto Excel no se encuentra en la ruta: {EXCEL_PATH}")

    mtime = os.path.getmtime(EXCEL_PATH)
    with _lock:
        if _excel_last_mtime and mtime == _excel_last_mtime:
            return _excel_cache

    log.info("[CACHE] Caché de Excel inválida o vacía. Cargando archivo físico...")

    wb = _load_workbook_safe()
    try:
        if 'Estimar' not in wb.sheetnames:
            raise ValueError("Error de integridad: La hoja 'Estimar' no existe en el estimador de presupuestos.")

        sheet_estimar = wb['Estimar']
        prom_start_row = _find_prom_row(sheet_estimar)

        parsed_data = {}
        for r in range(prom_start_row + 2, prom_start_row + 15):
            row_val = sheet_estimar.cell(r, 1).value
            if not row_val or row_val == 'Total general':
                continue
            mod = str(row_val).strip()
            exp_weeks, real_weeks, deploy_weeks, ttgl_weeks, asig_real = _parse_excel_row(r, sheet_estimar)
            parsed_data[mod] = {
                'explore_weeks': exp_weeks, 'realize_weeks': real_weeks,
                'deploy_weeks': deploy_weeks, 'ttgl_weeks': ttgl_weeks, 'fte_allocation': asig_real
            }

        with _lock:
            _excel_cache = parsed_data
            _excel_last_mtime = mtime
            log.info("[CACHE] Datos de Excel parseados e inyectados en la caché en memoria.")
            return _excel_cache
    finally:
        if wb:
            wb.close()

def _resolve_support_fraction(config, db_config, default_key):
    """Resuelve la fracción de soporte desde config o db_config, manejando conversión porcentaje->fracción."""
    support_raw = config.get('support_percentage')
    if support_raw is None:
        return db_config[default_key]
    value = float(support_raw)
    return value / 100.0 if value > 1.0 else value


def _validate_config(config, db_config, active_modules_list):
    """Valida y sobreescribe parámetros de configuración con restricciones de rango."""
    consulting_rate = float(config.get('consulting_rate') or db_config['tarifa_hora_consultor'])
    if not (10.0 <= consulting_rate <= 1000.0):
        raise ValueError("Límite financiero: La tarifa horaria de consultoría debe estar entre $10 y $1000 USD.")

    support_fraction = _resolve_support_fraction(config, db_config, 'porcentaje_ams')
    if not (0.0 <= support_fraction <= 1.0):
        raise ValueError("Límite financiero: El porcentaje de soporte AMS debe estar entre 0% y 100%.")

    saas_margin = db_config['margen_saas']
    if not (0.0 <= saas_margin <= 2.0):
        raise ValueError("Límite financiero: El margen SaaS debe estar entre 0% y 200%.")

    anos_roi = int(db_config['anos_roi'])
    if not (1 <= anos_roi <= 20):
        raise ValueError("Límite financiero: Las proyecciones de ROI deben estimarse entre 1 y 20 años.")

    tipo_cambio = float(config.get('exchange_rate') or db_config['tipo_cambio_pen'])
    if not (1.0 <= tipo_cambio <= 10.0):
        raise ValueError("Límite de localización: El tipo de cambio oficial debe estar en el rango de 1.0 a 10.0 PEN por USD.")

    factor_igv = float(config.get('igv_factor') or db_config['factor_igv'])
    if not (0.0 <= factor_igv <= 0.50):
        raise ValueError("Límite de localización: El factor de impuesto de IGV debe estar en el rango de 0% a 50%.")

    annual_revenue = float(config.get('annual_revenue') or 10000000.0)
    if annual_revenue <= 0.0:
        raise ValueError("Error comercial: La facturación anual de la empresa debe ser mayor a cero.")

    modular_licenses = DEFAULT_MODULAR_LICENSES.copy()
    for mod, val in config.get('modular_licenses', {}).items():
        if val is not None:
            modular_licenses[mod] = float(val)

    if not active_modules_list:
        raise ValueError("Error de configuración: Debe seleccionar al menos un módulo funcional de SAP.")

    return {
        'consulting_rate': consulting_rate, 'support_fraction': support_fraction,
        'saas_margin': saas_margin, 'anos_roi': anos_roi, 'tipo_cambio': tipo_cambio,
        'factor_igv': factor_igv, 'annual_revenue': annual_revenue, 'modular_licenses': modular_licenses
    }


def _process_module_details(active_modules_list, parsed_excel_data, modular_licenses, consulting_rate, saas_margin, tipo_cambio):
    """Procesa cada módulo activo calculando horas, costos y generando la estructura detallada."""
    modules_details = {}
    total_hours = 0.0
    ttgl_weeks_list = []

    for mod in active_modules_list:
        if mod in parsed_excel_data:
            md = parsed_excel_data[mod]
            mod_hours = md['ttgl_weeks'] * 40.0 * md['fte_allocation']
            lic_base = modular_licenses.get(mod, 15000.0)
            lic_final = lic_base * (1.0 + saas_margin)
            modules_details[mod] = {
                'explore_weeks': md['explore_weeks'], 'realize_weeks': md['realize_weeks'],
                'deploy_weeks': md['deploy_weeks'], 'ttgl_weeks': md['ttgl_weeks'],
                'fte_allocation': md['fte_allocation'], 'hours': mod_hours,
                'consulting_cost': mod_hours * consulting_rate,
                'consulting_cost_pen': mod_hours * consulting_rate * tipo_cambio,
                'saas_license_base': lic_base, 'saas_license_markup': lic_final,
                'saas_license_markup_pen': lic_final * tipo_cambio
            }
        else:
            lic_base = modular_licenses.get(mod, 15000.0)
            lic_final = lic_base * (1.0 + saas_margin)
            modules_details[mod] = {
                'explore_weeks': 4.0, 'realize_weeks': 10.0, 'deploy_weeks': 4.0,
                'ttgl_weeks': 18.0, 'fte_allocation': 0.50, 'hours': 360.0,
                'consulting_cost': 360.0 * consulting_rate,
                'consulting_cost_pen': 360.0 * consulting_rate * tipo_cambio,
                'saas_license_base': lic_base, 'saas_license_markup': lic_final,
                'saas_license_markup_pen': lic_final * tipo_cambio
            }
        total_hours += modules_details[mod]['hours']
        ttgl_weeks_list.append(modules_details[mod]['ttgl_weeks'])

    total_weeks = max(ttgl_weeks_list) if ttgl_weeks_list else 22.375
    return modules_details, total_hours, total_weeks


def _build_roi_projection(anos_roi, consulting_cost, licensing_cost_markup, support_cost_annual, savings_annual, tipo_cambio):
    """Genera la proyección de flujo de caja y ROI acumulado año a año."""
    roi_projection = []
    cum_tco = 0.0
    for y in range(1, anos_roi + 1):
        y_cost = (consulting_cost + licensing_cost_markup + support_cost_annual) if y == 1 else (licensing_cost_markup + support_cost_annual)
        cum_tco += y_cost
        cum_savings = savings_annual * y
        y_net = cum_savings - cum_tco
        y_roi = (y_net / cum_tco * 100.0) if cum_tco > 0 else 0.0
        roi_projection.append({
            'year': y, 'cost_annual': round(y_cost, 2), 'cost_annual_pen': round(y_cost * tipo_cambio, 2),
            'cum_tco': round(cum_tco, 2), 'cum_tco_pen': round(cum_tco * tipo_cambio, 2),
            'cum_savings': round(cum_savings, 2), 'cum_savings_pen': round(cum_savings * tipo_cambio, 2),
            'net_benefit': round(y_net, 2), 'net_benefit_pen': round(y_net * tipo_cambio, 2), 'roi_pct': round(y_roi, 1)
        })
    return roi_projection


def _build_bimoneda_summary(total_investment_y1_net, consulting_cost, licensing_cost_markup, support_cost_annual, savings_annual, tco_project_usd, tco_project_pen, factor_igv, tipo_cambio):
    """Construye resúmenes en USD y PEN con costos netos, IGV y facturables."""
    usd = {
        'net_investment': round(total_investment_y1_net, 2),
        'igv': round(total_investment_y1_net * factor_igv, 2),
        'total_facturable': round(total_investment_y1_net * (1.0 + factor_igv), 2),
        'net_investment_str': format_usd(total_investment_y1_net),
        'igv_str': format_usd(total_investment_y1_net * factor_igv),
        'total_facturable_str': format_usd(total_investment_y1_net * (1.0 + factor_igv)),
        'consulting_cost': round(consulting_cost, 2), 'consulting_cost_str': format_usd(consulting_cost),
        'licensing_cost': round(licensing_cost_markup, 2), 'licensing_cost_str': format_usd(licensing_cost_markup),
        'support_cost': round(support_cost_annual, 2), 'support_cost_str': format_usd(support_cost_annual),
        'savings_annual': round(savings_annual, 2), 'savings_annual_str': format_usd(savings_annual),
        'tco_project': round(tco_project_usd, 2), 'tco_project_str': format_usd(tco_project_usd)
    }
    pen = {
        'net_investment': round(total_investment_y1_net * tipo_cambio, 2),
        'igv': round(total_investment_y1_net * factor_igv * tipo_cambio, 2),
        'total_facturable': round(total_investment_y1_net * (1.0 + factor_igv) * tipo_cambio, 2),
        'net_investment_str': format_pen(total_investment_y1_net * tipo_cambio),
        'igv_str': format_pen(total_investment_y1_net * factor_igv * tipo_cambio),
        'total_facturable_str': format_pen(total_investment_y1_net * (1.0 + factor_igv) * tipo_cambio),
        'consulting_cost': round(consulting_cost * tipo_cambio, 2),
        'consulting_cost_str': format_pen(consulting_cost * tipo_cambio),
        'licensing_cost': round(licensing_cost_markup * tipo_cambio, 2),
        'licensing_cost_str': format_pen(licensing_cost_markup * tipo_cambio),
        'support_cost': round(support_cost_annual * tipo_cambio, 2),
        'support_cost_str': format_pen(support_cost_annual * tipo_cambio),
        'savings_annual': round(savings_annual * tipo_cambio, 2),
        'savings_annual_str': format_pen(savings_annual * tipo_cambio),
        'tco_project': round(tco_project_pen, 2), 'tco_project_str': format_pen(tco_project_pen)
    }
    return usd, pen


def calculate_financials(active_modules_list, config=None):
    """
    Realiza la lectura cacheada del estimador y computa de forma exacta las horas de consultoría,
    costos de licencias (SaaS) con recargo de margen, soporte de post-go-live (AMS), conversión multimoneda
    (USD/PEN), cálculo del IGV del 18%, ROI proyectado año a año y período de recupero.

    Parámetros:
    - active_modules_list (list): Lista de módulos SAP a considerar (ej: ['FI', 'MM', 'CO', 'PP'])
    - config (dict, opcional): Diccionario con parámetros manuales que sobreescriben la BD.

    Retorna:
    - dict: Contiene la estructura detallada por módulo y el resumen financiero bimoneda final.
    """
    db_config = load_db_config()
    if config is None:
        config = {}

    cfg = _validate_config(config, db_config, active_modules_list)
    parsed_excel_data = load_excel_data_cached()

    modules_details, total_hours, total_weeks = _process_module_details(
        active_modules_list, parsed_excel_data, cfg['modular_licenses'],
        cfg['consulting_rate'], cfg['saas_margin'], cfg['tipo_cambio']
    )

    consulting_cost = total_hours * cfg['consulting_rate']
    licensing_cost_base = sum(cfg['modular_licenses'].get(mod, 15000.0) for mod in active_modules_list)
    licensing_cost_markup = licensing_cost_base * (1.0 + cfg['saas_margin'])
    support_cost_annual = consulting_cost * cfg['support_fraction']
    total_investment_y1_net = consulting_cost + licensing_cost_markup + support_cost_annual

    savings_annual = cfg['annual_revenue'] * 0.015  # 1.5% savings estimate based on industry benchmarks
    tco_project_usd = consulting_cost + (licensing_cost_markup * cfg['anos_roi']) + (support_cost_annual * cfg['anos_roi'])
    tco_project_pen = tco_project_usd * cfg['tipo_cambio']
    savings_project_usd = savings_annual * cfg['anos_roi']
    net_savings_usd = savings_project_usd - tco_project_usd
    roi_project = (net_savings_usd / tco_project_usd * 100.0) if tco_project_usd > 0 else 0.0
    payback_period = (consulting_cost + licensing_cost_markup + support_cost_annual) / savings_annual if savings_annual > 0 else 0.0

    roi_projection = _build_roi_projection(
        cfg['anos_roi'], consulting_cost, licensing_cost_markup,
        support_cost_annual, savings_annual, cfg['tipo_cambio']
    )

    usd_summary, pen_summary = _build_bimoneda_summary(
        total_investment_y1_net, consulting_cost, licensing_cost_markup,
        support_cost_annual, savings_annual, tco_project_usd, tco_project_pen,
        cfg['factor_igv'], cfg['tipo_cambio']
    )

    return {
        'modules': modules_details,
        'summary': {
            'total_weeks': round(total_weeks, 2), 'total_hours': round(total_hours, 2),
            'consulting_cost': round(consulting_cost, 2),
            'licensing_cost': round(licensing_cost_markup, 2),
            'licensing_cost_base': round(licensing_cost_base, 2),
            'support_cost': round(support_cost_annual, 2),
            'total_investment': round(total_investment_y1_net * (1.0 + cfg['factor_igv']), 2),
            'total_investment_net': round(total_investment_y1_net, 2),
            'savings_annual': round(savings_annual, 2),
            'tco_five_years': round(tco_project_usd, 2),
            'savings_five_years': round(savings_project_usd, 2),
            'net_savings_five_years': round(net_savings_usd, 2),
            'roi_five_years': round(roi_project, 1),
            'payback_period': round(payback_period, 2),
            'anos_roi': cfg['anos_roi'], 'margen_saas': cfg['saas_margin'],
            'factor_igv': cfg['factor_igv'], 'tipo_cambio_pen': cfg['tipo_cambio'],
            'usd': usd_summary, 'pen': pen_summary, 'roi_projection': roi_projection
        }
    }

if __name__ == "__main__":
    # Test rápido de ejecución
    test_results = calculate_financials(['FI', 'CO', 'MM', 'SD'])
    print("\n--- Resultados de Prueba del Motor Financiero (Localización Peruana) ---")
    print("Inversión Inicial USD (Neta):", test_results['summary']['usd']['net_investment_str'])
    print("Inversión Inicial USD (IGV):", test_results['summary']['usd']['igv_str'])
    print("Inversión Inicial USD (Facturable):", test_results['summary']['usd']['total_facturable_str'])
    print("Inversión Inicial PEN (Facturable):", test_results['summary']['pen']['total_facturable_str'])
    print("ROI proyectado 5 años:", test_results['summary']['roi_five_years'], "%")
